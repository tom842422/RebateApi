import os
import openpyxl
import pprint
import os
from openpyxl import load_workbook
from read_yaml import ReadYaml as RY


class MyExcel:

    def __init__(self, filename, sheetname='sheet1'):
        """
        判断路径是否存在。如果不存在，抛异常。
        :param excel_path: 完整的excel文件路径
        """
        # self.filename = eval(RY('filename').get_data())
        # self.sheetname = eval(RY('sheetname').get_data())
        self.filename = filename
        self.sheetname = sheetname
        excel_path = os.path.join(
            os.path.join(
                os.path.dirname(os.path.dirname(os.path.abspath(__file__))) +
                '\\datas\\') + self.filename)

        if os.path.isfile(excel_path):
            if excel_path.endswith(".xlsx"):
                self.wb = load_workbook(excel_path)
            else:
                print("文件不是以xlsx结尾,不支持处理。")
        else:
            print("文件路径不存在。")
            raise FileNotFoundError
        if self.sheetname not in self.wb.sheetnames:
            print("表单名称不存在！")
            self.sh = None
        else:
            self.sh = self.wb[self.sheetname]

    def read_all_data(self):
        if self.sh:
            cases_dict = {}
            # 获取表单里的所有数据
            all_values = list(self.sh.values)
            # 获取表单第一行，作为key
            keys = all_values[0]
            # 遍历表单，从第2行开始
            for values in all_values[1:]:
                # 把key和每一行进行组合成字典
                case = dict(zip(keys, values))
                cases_dict |= case
            return cases_dict

    def read_column(self):
        return self.wb.columns


if __name__ == '__main__':
    # res = MyExcel(filename='买家订单明细.xlsx').read_all_data()
    res = MyExcel(filename='买家订单明细.xlsx').read_column()
    print(res)
    print(type(res))
    # print(res['商品名称'])
